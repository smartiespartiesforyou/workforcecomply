from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import os
import re
import zipfile
import shutil
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from PyPDF2 import PdfMerger

from oig_screenshot import capture_oig, close_oig_session
from cna_screenshot import capture_cna, close_cna_session
from adverse_screenshot import capture_adverse, close_adverse_session

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

UPLOAD_FOLDER = "uploads"
RUNS_FOLDER = "runs"
BACKEND_BASE_URL = "https://workforcecomply-backend-docker.onrender.com"

COMBINED_WORKERS = 3

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RUNS_FOLDER, exist_ok=True)


def safe_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_ssn(value):
    return re.sub(r"\D", "", safe_text(value))


def cleanup_old_runs(folder, days=2):
    now = datetime.now()
    for name in os.listdir(folder):
        path = os.path.join(folder, name)
        if os.path.isdir(path):
            try:
                created = datetime.fromtimestamp(os.path.getctime(path))
                if now - created > timedelta(days=days):
                    shutil.rmtree(path)
            except Exception:
                pass


def merge_pdfs(pdf_paths, output_path):
    valid_paths = [p for p in pdf_paths if p and os.path.exists(p)]
    if not valid_paths:
        return None

    merger = PdfMerger()
    for pdf_path in valid_paths:
        merger.append(pdf_path)
    merger.write(output_path)
    merger.close()
    return output_path


def normalize_status(issue_list, check_name):
    relevant = [issue for issue in issue_list if check_name in issue.upper()]

    if not relevant:
        return "Clear"

    text = " | ".join(relevant).upper()

    if "INVALID SSN" in text:
        return "Invalid SSN"
    if "MATCH" in text:
        return "Match Found"
    if "FOUND" in text:
        return "Match Found"
    if "NOT ACTIVE" in text:
        return "Not Active"
    if "PROOF MISSING" in text:
        return "Proof Missing"
    if "ERROR" in text:
        return "Error"
    if "REVIEW" in text:
        return "Review Needed"

    return "Review Needed"


def create_results_excel(employee_results, output_path, mode="combined"):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    flagged_rows = []

    for e in employee_results:
        if e["flagged"]:
            if mode == "oig":
                flagged_rows.append({
                    "First Name": e["First Name"],
                    "Last Name": e["Last Name"],
                    "SSN": e["SSN"],
                    "OIG": normalize_status(e["issues"], "OIG"),
                    "Status": "Attention Required"
                })
            elif mode == "cna":
                flagged_rows.append({
                    "First Name": e["First Name"],
                    "Last Name": e["Last Name"],
                    "SSN": e["SSN"],
                    "CNA": normalize_status(e["issues"], "CNA"),
                    "Status": "Attention Required"
                })
            elif mode == "adverse":
                flagged_rows.append({
                    "First Name": e["First Name"],
                    "Last Name": e["Last Name"],
                    "SSN": e["SSN"],
                    "Adverse": normalize_status(e["issues"], "ADVERSE"),
                    "Status": "Attention Required"
                })
            else:
                flagged_rows.append({
                    "First Name": e["First Name"],
                    "Last Name": e["Last Name"],
                    "SSN": e["SSN"],
                    "OIG": normalize_status(e["issues"], "OIG"),
                    "CNA": normalize_status(e["issues"], "CNA"),
                    "Adverse": normalize_status(e["issues"], "ADVERSE"),
                    "Status": "Attention Required"
                })

    if flagged_rows:
        df = pd.DataFrame(flagged_rows)
    else:
        if mode == "oig":
            df = pd.DataFrame(columns=[
                "First Name", "Last Name", "SSN", "OIG", "Status"
            ])
        elif mode == "cna":
            df = pd.DataFrame(columns=[
                "First Name", "Last Name", "SSN", "CNA", "Status"
            ])
        elif mode == "adverse":
            df = pd.DataFrame(columns=[
                "First Name", "Last Name", "SSN", "Adverse", "Status"
            ])
        else:
            df = pd.DataFrame(columns=[
                "First Name", "Last Name", "SSN", "OIG", "CNA", "Adverse", "Status"
            ])

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Flagged Employees"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)


def build_zip(run_folder, zip_path, include_folders):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        results_path = os.path.join(run_folder, "Results.xlsx")
        if os.path.exists(results_path):
            z.write(results_path, "Results.xlsx")

        for folder_name in include_folders:
            folder_path = os.path.join(run_folder, folder_name)
            if not os.path.exists(folder_path):
                continue

            for root, dirs, files in os.walk(folder_path):
                for file_name in files:
                    full_path = os.path.join(root, file_name)
                    relative_path = os.path.relpath(full_path, run_folder)
                    z.write(full_path, relative_path)


def run_oig_safe(first, last, save_folder):
    return capture_oig(first, last, save_folder)


def run_cna_safe(ssn, save_folder):
    return capture_cna(ssn, save_folder)


def run_adverse_safe(first, last, ssn, save_folder):
    return capture_adverse(first, last, ssn, save_folder)


def prepare_upload(file, run_id):
    upload_name = os.path.basename(file.filename)
    upload_path = os.path.join(UPLOAD_FOLDER, f"{run_id}_{upload_name}")
    file.save(upload_path)
    return upload_path


def read_input_dataframe(upload_path):
    df = pd.read_excel(upload_path)

    required_columns = ["First Name", "Last Name", "SSN"]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    return df


def process_combined_run(df, run_folder):
    oig_folder = os.path.join(run_folder, "OIG_Report")
    cna_folder = os.path.join(run_folder, "CNA_Report")
    adverse_folder = os.path.join(run_folder, "Adverse_Actions_Report")

    os.makedirs(oig_folder, exist_ok=True)
    os.makedirs(cna_folder, exist_ok=True)
    os.makedirs(adverse_folder, exist_ok=True)

    employee_results = []
    oig_paths = []
    cna_paths = []
    adverse_paths = []

    try:    
        def process_employee(row):
            first = safe_text(row["First Name"])
            last = safe_text(row["Last Name"])
            ssn_raw = row["SSN"]
            ssn = clean_ssn(ssn_raw)

            employee = {
                "First Name": first,
                "Last Name": last,
                "SSN": ssn,
                "issues": [],
                "flagged": False
            }

            try:
                oig_result = run_oig_safe(first, last, oig_folder)
                oig_pdf = oig_result.get("pdf_path")

                if oig_pdf:
                    oig_paths.append(oig_pdf)
                else:
                    error = oig_result.get("error")
                    if error:
                        employee["issues"].append(f"REVIEW NEEDED - OIG ERROR: {error}")
                    else:
                        employee["issues"].append("REVIEW NEEDED - OIG PROOF MISSING")
            except Exception as e:
                employee["issues"].append(f"REVIEW NEEDED - OIG ERROR: {str(e)}")

            if len(ssn) != 9:
                employee["issues"].append("ERROR - INVALID SSN FOR CNA")
            else:
                try:
                    cna_result = run_cna_safe(ssn, cna_folder)
                    cna_pdf = cna_result.get("pdf_path")
                    cna_status = cna_result.get("cna_result", "")

                    if cna_pdf:
                        cna_paths.append(cna_pdf)

                    if cna_status == "not_active":
                        employee["issues"].append("CNA NOT ACTIVE")
                    elif cna_status == "review_needed":
                        employee["issues"].append("REVIEW NEEDED - CNA REVIEW")
                    elif cna_status == "error":
                        error = cna_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - CNA ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - CNA ERROR")

                except Exception as e:
                    employee["issues"].append(f"REVIEW NEEDED - CNA ERROR: {str(e)}")

            if len(ssn) != 9:
                employee["issues"].append("ERROR - INVALID SSN FOR ADVERSE")
            else:
                try:
                    adverse_result = run_adverse_safe(
                        first,
                        last,
                        ssn,
                        adverse_folder
                    )
                    adverse_pdf = adverse_result.get("pdf_path")

                    if adverse_pdf:
                        adverse_paths.append(adverse_pdf)

                    adverse_status = safe_text(
                        adverse_result.get("adverse_result", adverse_result.get("status", ""))
                    ).lower()

                    if adverse_status in ("match", "found", "review_needed"):
                        detail = safe_text(adverse_result.get("detail")) or "ADVERSE ACTION FOUND"
                        employee["issues"].append(f"REVIEW NEEDED - ADVERSE: {detail}")
                    elif adverse_status == "error":
                        error = adverse_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - ADVERSE ERROR")
                    elif not adverse_pdf:
                        error = adverse_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - ADVERSE PROOF MISSING")

                except Exception as e:
                    employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {str(e)}")

            if employee["issues"]:
                employee["flagged"] = True

            return employee

        with ThreadPoolExecutor(max_workers=COMBINED_WORKERS) as executor:
            employee_results.extend(
                list(executor.map(process_employee, [row for _, row in df.iterrows()]))
            )

    finally:
        close_oig_session()
        close_cna_session()
        close_adverse_session()

    merge_pdfs(oig_paths, os.path.join(oig_folder, "OIG_Merged.pdf"))
    merge_pdfs(cna_paths, os.path.join(cna_folder, "CNA_Merged.pdf"))
    merge_pdfs(adverse_paths, os.path.join(adverse_folder, "Adverse_Actions_Merged.pdf"))

    results_excel_path = os.path.join(run_folder, "Results.xlsx")
    create_results_excel(employee_results, results_excel_path, mode="combined")

    zip_path = os.path.join(run_folder, "output.zip")
    build_zip(
        run_folder,
        zip_path,
        include_folders=["OIG_Report", "CNA_Report", "Adverse_Actions_Report"]
    )

    return employee_results


def process_oig_only_run(df, run_folder):
    oig_folder = os.path.join(run_folder, "OIG_Report")
    os.makedirs(oig_folder, exist_ok=True)

    employee_results = []
    oig_paths = []

    try:
        for _, row in df.iterrows():
            first = safe_text(row["First Name"])
            last = safe_text(row["Last Name"])
            ssn_raw = row["SSN"]
            ssn = clean_ssn(ssn_raw)

            employee = {
                "First Name": first,
                "Last Name": last,
                "SSN": ssn,
                "issues": [],
                "flagged": False
            }

            try:
                oig_result = run_oig_safe(first, last, oig_folder)
                oig_pdf = oig_result.get("pdf_path")

                if oig_pdf:
                    oig_paths.append(oig_pdf)
                else:
                    error = oig_result.get("error")
                    if error:
                        employee["issues"].append(f"REVIEW NEEDED - OIG ERROR: {error}")
                    else:
                        employee["issues"].append("REVIEW NEEDED - OIG PROOF MISSING")
            except Exception as e:
                employee["issues"].append(f"REVIEW NEEDED - OIG ERROR: {str(e)}")

            if employee["issues"]:
                employee["flagged"] = True

            employee_results.append(employee)

    finally:
        close_oig_session()

    merge_pdfs(oig_paths, os.path.join(oig_folder, "OIG_Merged.pdf"))

    results_excel_path = os.path.join(run_folder, "OIG_Results.xlsx")
    create_results_excel(employee_results, results_excel_path, mode="oig")

    zip_path = os.path.join(run_folder, "OIG_Report.zip")
    build_zip(run_folder, zip_path, include_folders=["OIG_Report"])

    return employee_results


def process_cna_only_run(df, run_folder):
    cna_folder = os.path.join(run_folder, "CNA_Report")
    os.makedirs(cna_folder, exist_ok=True)

    employee_results = []
    cna_paths = []

    try:
        for _, row in df.iterrows():
            first = safe_text(row["First Name"])
            last = safe_text(row["Last Name"])
            ssn_raw = row["SSN"]
            ssn = clean_ssn(ssn_raw)

            employee = {
                "First Name": first,
                "Last Name": last,
                "SSN": ssn,
                "issues": [],
                "flagged": False
            }

            if len(ssn) != 9:
                employee["issues"].append("ERROR - INVALID SSN FOR CNA")
            else:
                try:
                    cna_result = run_cna_safe(ssn, cna_folder)
                    cna_pdf = cna_result.get("pdf_path")
                    cna_status = cna_result.get("cna_result", "")

                    if cna_pdf:
                        cna_paths.append(cna_pdf)

                    if cna_status == "not_active":
                        employee["issues"].append("CNA NOT ACTIVE")
                    elif cna_status == "review_needed":
                        employee["issues"].append("REVIEW NEEDED - CNA REVIEW")
                    elif cna_status == "error":
                        error = cna_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - CNA ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - CNA ERROR")

                except Exception as e:
                    employee["issues"].append(f"REVIEW NEEDED - CNA ERROR: {str(e)}")

            if employee["issues"]:
                employee["flagged"] = True

            employee_results.append(employee)

    finally:
        close_cna_session()

    merge_pdfs(cna_paths, os.path.join(cna_folder, "CNA_Merged.pdf"))

    results_excel_path = os.path.join(run_folder, "CNA_Results.xlsx")
    create_results_excel(employee_results, results_excel_path, mode="cna")

    zip_path = os.path.join(run_folder, "CNA_Report.zip")
    build_zip(run_folder, zip_path, include_folders=["CNA_Report"])

    return employee_results


def process_adverse_only_run(df, run_folder):
    adverse_folder = os.path.join(run_folder, "Adverse_Actions_Report")
    os.makedirs(adverse_folder, exist_ok=True)

    employee_results = []
    adverse_paths = []

    try:
        for _, row in df.iterrows():
            first = safe_text(row["First Name"])
            last = safe_text(row["Last Name"])
            ssn_raw = row["SSN"]
            ssn = clean_ssn(ssn_raw)

            employee = {
                "First Name": first,
                "Last Name": last,
                "SSN": ssn,
                "issues": [],
                "flagged": False
            }

            if len(ssn) != 9:
                employee["issues"].append("ERROR - INVALID SSN FOR ADVERSE")
            else:
                try:
                    adverse_result = run_adverse_safe(first, last, ssn, adverse_folder)
                    adverse_pdf = adverse_result.get("pdf_path")

                    if adverse_pdf:
                        adverse_paths.append(adverse_pdf)

                    adverse_status = safe_text(
                        adverse_result.get("adverse_result", adverse_result.get("status", ""))
                    ).lower()

                    if adverse_status in ("match", "found", "review_needed"):
                        detail = safe_text(adverse_result.get("detail")) or "ADVERSE ACTION FOUND"
                        employee["issues"].append(f"REVIEW NEEDED - ADVERSE: {detail}")
                    elif adverse_status == "error":
                        error = adverse_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - ADVERSE ERROR")
                    elif not adverse_pdf:
                        error = adverse_result.get("error")
                        if error:
                            employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {error}")
                        else:
                            employee["issues"].append("REVIEW NEEDED - ADVERSE PROOF MISSING")

                except Exception as e:
                    employee["issues"].append(f"REVIEW NEEDED - ADVERSE ERROR: {str(e)}")

            if employee["issues"]:
                employee["flagged"] = True

            employee_results.append(employee)

    finally:
        close_adverse_session()

    merge_pdfs(adverse_paths, os.path.join(adverse_folder, "Adverse_Actions_Merged.pdf"))

    results_excel_path = os.path.join(run_folder, "Results.xlsx")
    create_results_excel(employee_results, results_excel_path, mode="adverse")

    zip_path = os.path.join(run_folder, "output.zip")
    build_zip(run_folder, zip_path, include_folders=["Adverse_Actions_Report"])

    return employee_results


def make_response(run_id, employee_results):
    total = len(employee_results)
    flagged = sum(1 for e in employee_results if e["flagged"])
    clear = total - flagged

    return jsonify({
        "summary": {
            "total_employees": total,
            "clear_count": clear,
            "attention_needed": flagged
        },
        "downloads": {
            "combined_pdf_url": f"{BACKEND_BASE_URL}/api/download/{run_id}/zip",
            "individual_zip_url": f"{BACKEND_BASE_URL}/api/download/{run_id}/zip",
            "results_excel_url": f"{BACKEND_BASE_URL}/api/download/{run_id}/results-excel"
        }
    })


@app.route("/", methods=["GET"])
def home():
    return jsonify({"status": "ok", "message": "WorkforceComply backend is running"})


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/api/run-checks", methods=["POST"])
def run_checks():
    cleanup_old_runs(RUNS_FOLDER)

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    os.makedirs(run_folder, exist_ok=True)

    upload_path = prepare_upload(file, run_id)

    try:
        df = read_input_dataframe(upload_path)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    employee_results = process_combined_run(df, run_folder)
    return make_response(run_id, employee_results)


@app.route("/api/run-oig", methods=["POST"])
def run_oig_only():
    cleanup_old_runs(RUNS_FOLDER)

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    os.makedirs(run_folder, exist_ok=True)

    upload_path = prepare_upload(file, run_id)

    try:
        df = read_input_dataframe(upload_path)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    employee_results = process_oig_only_run(df, run_folder)
    return make_response(run_id, employee_results)


@app.route("/api/run-cna", methods=["POST"])
def run_cna_only():
    cleanup_old_runs(RUNS_FOLDER)

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    os.makedirs(run_folder, exist_ok=True)

    upload_path = prepare_upload(file, run_id)

    try:
        df = read_input_dataframe(upload_path)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    employee_results = process_cna_only_run(df, run_folder)
    return make_response(run_id, employee_results)


@app.route("/api/run-adverse", methods=["POST"])
def run_adverse_only():
    cleanup_old_runs(RUNS_FOLDER)

    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    os.makedirs(run_folder, exist_ok=True)

    upload_path = prepare_upload(file, run_id)

    try:
        df = read_input_dataframe(upload_path)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    employee_results = process_adverse_only_run(df, run_folder)
    return make_response(run_id, employee_results)


@app.route("/api/download/<run_id>/zip", methods=["GET"])
def download_zip(run_id):
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    zip_path = None

    if not os.path.exists(run_folder):
        return jsonify({"error": "Run folder not found"}), 404

    for f in os.listdir(run_folder):
        if f.endswith(".zip"):
            zip_path = os.path.join(run_folder, f)
            break

    if not zip_path or not os.path.exists(zip_path):
        return jsonify({"error": "ZIP file not found"}), 404

    return send_file(zip_path, as_attachment=True)


@app.route("/api/download/<run_id>/results-excel", methods=["GET"])
def download_results_excel(run_id):
    run_folder = os.path.join(RUNS_FOLDER, run_id)
    results_excel_path = None

    for f in os.listdir(run_folder):
        if f.endswith(".xlsx"):
            results_excel_path = os.path.join(run_folder, f)
            break

    if not results_excel_path or not os.path.exists(results_excel_path):
        return jsonify({"error": "Results Excel file not found"}), 404

    return send_file(results_excel_path, as_attachment=True)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
